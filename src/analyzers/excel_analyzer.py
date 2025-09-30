"""
Excel analyzer for .xlsx and .xls files
"""

import logging
from pathlib import Path
from typing import Dict, Any

try:
    import openpyxl
    import pandas as pd
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from . import BaseAnalyzer

class ExcelAnalyzer(BaseAnalyzer):
    """Analyzer for Excel files"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        if not EXCEL_AVAILABLE:
            self.logger.warning("openpyxl/pandas not available. Excel analysis will be limited.")
    
    def extract_text(self, file_path: Path) -> str:
        """Extract text content from Excel files"""
        if not EXCEL_AVAILABLE:
            return "Excel analysis requires openpyxl and pandas packages"
        
        try:
            # Read all sheets
            excel_file = pd.ExcelFile(file_path)
            text_content = []
            
            for sheet_name in excel_file.sheet_names:
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    
                    text_content.append(f"\n--- Sheet: {sheet_name} ---")
                    text_content.append(f"Dimensions: {df.shape[0]} rows x {df.shape[1]} columns")
                    
                    if not df.empty:
                        # Add column headers
                        text_content.append("Columns: " + ", ".join(str(col) for col in df.columns))
                        
                        # Add first few rows as preview
                        preview_rows = min(5, len(df))
                        if preview_rows > 0:
                            text_content.append("Sample data:")
                            for idx, row in df.head(preview_rows).iterrows():
                                row_text = " | ".join(str(val) for val in row.values if pd.notna(val))
                                if row_text.strip():
                                    text_content.append(f"  {row_text}")
                    
                except Exception as e:
                    text_content.append(f"Error reading sheet {sheet_name}: {str(e)}")
            
            return "\n".join(text_content)
            
        except Exception as e:
            self.logger.error(f"Error reading Excel file {file_path}: {e}")
            return f"Error reading Excel file: {str(e)}"
    
    def extract_metadata(self, file_path: Path) -> Dict[str, Any]:
        """Extract metadata from Excel files"""
        if not EXCEL_AVAILABLE:
            return {'error': 'openpyxl/pandas not available'}
        
        try:
            # Use openpyxl for detailed metadata
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            
            metadata = {
                'file_type': 'excel',
                'sheet_count': len(workbook.sheetnames),
                'sheet_names': workbook.sheetnames,
            }
            
            # Get workbook properties
            props = workbook.properties
            if props.title:
                metadata['title'] = props.title
            if props.creator:
                metadata['creator'] = props.creator
            if props.description:
                metadata['description'] = props.description
            if props.subject:
                metadata['subject'] = props.subject
            if props.keywords:
                metadata['keywords'] = props.keywords
            if props.created:
                metadata['created'] = str(props.created)
            if props.modified:
                metadata['modified'] = str(props.modified)
            if props.lastModifiedBy:
                metadata['last_modified_by'] = props.lastModifiedBy
            
            # Analyze each sheet
            sheet_info = {}
            for sheet_name in workbook.sheetnames:
                try:
                    sheet = workbook[sheet_name]
                    
                    # Count rows and columns with data
                    max_row = sheet.max_row
                    max_col = sheet.max_column
                    
                    # Try to get actual data dimensions
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    
                    sheet_info[sheet_name] = {
                        'max_row': max_row,
                        'max_column': max_col,
                        'data_rows': len(df),
                        'data_columns': len(df.columns),
                        'has_data': not df.empty,
                        'column_names': list(df.columns) if not df.empty else []
                    }
                    
                except Exception as e:
                    sheet_info[sheet_name] = {'error': str(e)}
            
            metadata['sheets'] = sheet_info
            workbook.close()
            
            return metadata
            
        except Exception as e:
            self.logger.error(f"Error extracting Excel metadata from {file_path}: {e}")
            return {'error': str(e)}